# import selenium
# import time
import openpyxl

from openpyxl import Workbook
import selenium
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options

date_c=0
chrome_options = Options()
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--headless")
chrome_options.headless = True
driver = webdriver.Chrome('chromedriver.exe',options=chrome_options)
a=1
b=1

with open('maa.txt','r') as maa:
	datecc='//*[@id="stat_table"]/tbody/tr['

	for line in maa:
		m=line
		driver.get(m)
		date_c=0
		click1=driver.find_element_by_xpath("//div[text()='Показать все дни']")
		ActionChains(driver).click(click1).perform()
		lnin=str(m[30:40])+'.xlsx'
		lnin=lnin.replace('/','')
		wb=Workbook()
		ws=wb.active
		# ws1=wb.create_sheet('Sheet')
		wb.save(lnin)
		# wb = Workbook(lnin)
		# ws = wb.active
		# ws1 = wb.create_sheet('Sheet')
		# wb.save(lnin)
		for i in range(500):
			date_c+=1
			
			a+=1
			b+=1

		# Это не дата нихуя это количество зараженных я даун нахуя написал дате когда нихуя не дате
			try:
				
				
				date=driver.find_element_by_xpath(datecc+str(date_c)+']/td[2]/b').text


				A11='a'+str(a)

				worksheet=wb['Sheet']
				worksheet[A11]=str(date)
				wb.save(lnin)

			except selenium.common.exceptions.NoSuchElementException:
				pass
		# Вот тут будет дата ебана и какую переменную сувать ну бля нахуй сука
			try:
				
				bebra=driver.find_element_by_xpath(datecc+str(date_c)+']/td[1]').text
				B11='b'+str(b)
				worksheet=wb['Sheet']
				worksheet[B11]=str(bebra)
				wb.save(lnin)
			except selenium.common.exceptions.NoSuchElementException:
				pass
			print(bebra)
		
		# line=line[30:50]
		# print(line)	
# lnin='1213.xlsx'

# wb = Workbook(lnin)
# ws = wb.active
# ws1 = wb.create_sheet('Sheet')
# wb.save(lnin)